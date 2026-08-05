"""Epic 10 Story 10.7: tests for worker/executor/test_runner/report.py --
document-id expansion (bundled PERF, not-implemented RUN ids), the real
aggregate counts (POST=41 not the source doc's stale 33, Total=71 not 63),
TestStatus->Pass/Fail mapping, override precedence, and the TC-RUN-013
per-OSD sub-table. No DB/HTTP involved -- pure functions only.
"""

from worker.executor.test_runner import report
from worker.executor.test_runner.registry import TEST_CASES_BY_ID


def test_expand_document_ids_bundles_perf_001_003():
    assert report.expand_document_ids("TC-PERF-001-003") == ["TC-PERF-001", "TC-PERF-002", "TC-PERF-003"]


def test_expand_document_ids_identity_for_everything_else():
    assert report.expand_document_ids("TC-POST-010") == ["TC-POST-010"]
    assert report.expand_document_ids("TC-RUN-001") == ["TC-RUN-001"]


def test_map_status_to_pass_fail_all_states():
    assert report.map_status_to_pass_fail("pass", False) == "Pass"
    assert report.map_status_to_pass_fail("fail", False) == "Fail"
    assert report.map_status_to_pass_fail("error", False) == "Blocked"
    assert report.map_status_to_pass_fail("skip", False) == "N/A"
    assert report.map_status_to_pass_fail("running", False) == ""
    assert report.map_status_to_pass_fail("pending", False) == ""
    assert report.map_status_to_pass_fail(None, False) == ""


def test_build_report_rows_covers_all_71_real_document_ids():
    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    ids = [r.document_id for r in rows]
    assert len(ids) == len(set(ids))  # no duplicate document ids
    assert len(rows) == 71
    for not_impl in report.NOT_IMPLEMENTED_DOCUMENT_IDS:
        assert not_impl in ids
    assert "TC-PERF-001" in ids and "TC-PERF-002" in ids and "TC-PERF-003" in ids
    assert "TC-PERF-001-003" not in ids  # engine id itself must not leak into the document-id rows


def test_build_report_rows_not_implemented_ids_have_no_result():
    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    by_id = {r.document_id: r for r in rows}
    for doc_id in report.NOT_IMPLEMENTED_DOCUMENT_IDS:
        row = by_id[doc_id]
        assert row.pass_fail == ""
        assert row.priority is None
        assert "gap" in row.ghi_chu.lower() or "tài liệu" in row.ghi_chu


def test_build_report_rows_bundled_perf_rows_share_same_result():
    run_states = {
        "TC-PERF-001-003": {
            "status": "pass",
            "criteria": [{"description": "x", "passed": True, "detail": ""}],
            "notes": "",
            "finished_at": "2026-08-05T10:00:00",
        }
    }
    rows = report.build_report_rows(run_states, TEST_CASES_BY_ID, "tester")
    by_id = {r.document_id: r for r in rows}
    for doc_id in ("TC-PERF-001", "TC-PERF-002", "TC-PERF-003"):
        assert by_id[doc_id].pass_fail == "Pass"
        assert by_id[doc_id].ngay_thuc_hien == "2026-08-05"


def test_build_report_rows_never_run_is_blank_not_fail():
    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    by_id = {r.document_id: r for r in rows}
    assert by_id["TC-POST-001"].pass_fail == ""
    assert by_id["TC-POST-001"].ket_qua_thuc_te == "Chưa chạy"


def test_build_report_rows_error_maps_to_blocked_and_skip_to_na():
    run_states = {
        "TC-POST-001": {"status": "error", "criteria": [], "notes": "SSH timeout"},
        "TC-COMPAT-005": {"status": "skip", "criteria": [], "notes": "no OpenStack config"},
    }
    rows = report.build_report_rows(run_states, TEST_CASES_BY_ID, "tester")
    by_id = {r.document_id: r for r in rows}
    assert by_id["TC-POST-001"].pass_fail == "Blocked"
    assert by_id["TC-COMPAT-005"].pass_fail == "N/A"


def test_build_report_rows_override_prefixes_ghi_chu():
    run_states = {
        "TC-POST-002": {
            "status": "fail",
            "criteria": [],
            "notes": "",
            "overridden": True,
            "override_note": "confirmed healthy by hand",
        }
    }
    rows = report.build_report_rows(run_states, TEST_CASES_BY_ID, "tester")
    row = next(r for r in rows if r.document_id == "TC-POST-002")
    assert row.pass_fail == "Fail"
    assert row.ghi_chu == "[Override] confirmed healthy by hand"


def test_build_aggregate_table_real_counts_not_source_doc_stale_numbers():
    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    aggregate = report.build_aggregate_table(rows)
    by_group = {a["group"]: a for a in aggregate}
    assert by_group["RUN"]["tong_so_tc"] == 13
    assert by_group["POST"]["tong_so_tc"] == 41  # not the source document's stale 33
    assert by_group["COMPAT"]["tong_so_tc"] == 8
    assert by_group["PERF"]["tong_so_tc"] == 9
    assert by_group["Tổng"]["tong_so_tc"] == 71  # not the source document's stale 63


def test_build_aggregate_table_pass_fail_blocked_na_counts():
    run_states = {
        "TC-COMPAT-001": {"status": "pass", "criteria": []},
        "TC-COMPAT-002": {"status": "fail", "criteria": []},
        "TC-COMPAT-003": {"status": "error", "criteria": []},
        "TC-COMPAT-005": {"status": "skip", "criteria": []},
    }
    rows = report.build_report_rows(run_states, TEST_CASES_BY_ID, "tester")
    aggregate = report.build_aggregate_table(rows)
    compat = next(a for a in aggregate if a["group"] == "COMPAT")
    assert compat["pass"] == 1
    assert compat["fail"] == 1
    assert compat["blocked"] == 1
    assert compat["na"] == 1
    assert compat["tong_so_tc"] == 8


def test_build_run013_osd_table_none_without_background_state():
    assert report.build_run013_osd_table({}) is None
    assert report.build_run013_osd_table({"TC-RUN-013": {"status": "pending"}}) is None
    assert report.build_run013_osd_table({"TC-RUN-013": {"status": "pending", "background_state": None}}) is None


def test_build_run013_osd_table_reads_completed_list():
    run_states = {
        "TC-RUN-013": {
            "status": "pending",
            "background_state": {
                "completed": [
                    {"osd_id": 3, "seconds": 42.456, "exit_code": 0, "over_estimate": False},
                    {"osd_id": 7, "seconds": 900.1, "exit_code": 1, "over_estimate": True},
                ]
            },
        }
    }
    table = report.build_run013_osd_table(run_states)
    assert table == [
        {"osd_id": 3, "seconds": 42.5, "exit_code": 0, "over_estimate": False},
        {"osd_id": 7, "seconds": 900.1, "exit_code": 1, "over_estimate": True},
    ]


def test_build_exit_criteria_checklist_has_9_items_matching_source_doc():
    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    checklist = report.build_exit_criteria_checklist(rows)
    assert len(checklist) == 9
    assert all(isinstance(item, str) and isinstance(checked, bool) for item, checked, _ in checklist)


def test_build_exit_criteria_checklist_defect_and_perf_items_never_auto_checked():
    run_states = {tc_id: {"status": "pass", "criteria": []} for tc_id in TEST_CASES_BY_ID}
    rows = report.build_report_rows(run_states, TEST_CASES_BY_ID, "tester")
    checklist = report.build_exit_criteria_checklist(rows)
    # item 2 (defect) and item 6 (perf regression) always unchecked -- no defect tracker,
    # and PERF criteria are passed=None by design so status can never actually be "pass" for
    # real, but guard the invariant explicitly regardless of run_states content.
    assert checklist[1][1] is False
    assert checklist[5][1] is False


def test_build_exit_criteria_checklist_item1_all_p1_pass_and_p2_threshold():
    run_states = {tc_id: {"status": "pass", "criteria": []} for tc_id in TEST_CASES_BY_ID}
    rows = report.build_report_rows(run_states, TEST_CASES_BY_ID, "tester")
    checklist = report.build_exit_criteria_checklist(rows)
    assert checklist[0][1] is True


def test_build_markdown_report_contains_all_sections():
    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    aggregate = report.build_aggregate_table(rows)
    checklist = report.build_exit_criteria_checklist(rows)
    markdown = report.build_markdown_report("tester", rows, aggregate, None, checklist)
    assert "TC-RUN-001" in markdown
    assert "TC-PERF-002" in markdown
    assert "Bảng tổng hợp" in markdown
    assert "Tiêu chí kết thúc" in markdown
    assert "Chi tiết TC-RUN-013" not in markdown  # no run013_table -> no sub-table


def test_build_markdown_report_includes_run013_subtable_when_present():
    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    aggregate = report.build_aggregate_table(rows)
    checklist = report.build_exit_criteria_checklist(rows)
    run013_table = [{"osd_id": 3, "seconds": 42.5, "exit_code": 0, "over_estimate": False}]
    markdown = report.build_markdown_report("tester", rows, aggregate, run013_table, checklist)
    assert "Chi tiết TC-RUN-013" in markdown
    assert "| 3 | 42.5 | 0 | Không |" in markdown


def test_build_excel_workbook_round_trips_via_openpyxl():
    import io

    from openpyxl import load_workbook

    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    aggregate = report.build_aggregate_table(rows)
    xlsx_bytes = report.build_excel_workbook("tester", rows, aggregate)

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Ket qua test case", "Tong hop"]
    ws1 = wb["Ket qua test case"]
    assert ws1.cell(row=1, column=1).value == "Test Case ID"
    assert ws1.max_row == len(rows) + 1  # header + 71 rows
    ws2 = wb["Tong hop"]
    assert ws2.max_row == len(aggregate) + 1


def test_build_copy_summary_text_has_totals_and_table():
    rows = report.build_report_rows({}, TEST_CASES_BY_ID, "tester")
    aggregate = report.build_aggregate_table(rows)
    summary = report.build_copy_summary_text(rows, aggregate)
    assert "PASS" in summary
    assert "P1:" in summary
    assert "Tổng" in summary
