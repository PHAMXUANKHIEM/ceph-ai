"""Route tests for Epic 10 Story 10.2 -- dashboard/routes/test_runner.py's
4 new config endpoints (GET/POST /api/test-runner/config, POST .../config/
baseline/{baseline_key}, POST .../ssh-check). Follows this project's
existing FastAPI route-testing conventions (see test_dashboard_backups.py,
test_dashboard_upgrade_procedure.py): the `dashboard_client` fixture
(isolated in-memory DB + fixed test credentials) and a `_login` helper that
posts to /login before each authenticated request.

Story 10.6 (below, from "-- Story 10.6" section) adds tests for the 4 new
run-engine endpoints (GET /tests, POST .../run, GET .../result, POST .../
override). Those tests fake the whole test-case registry (TEST_CASES_BY_ID)
rather than exercising a real Group A-D TestCase over SSH -- real classes
would need actual network access and the default 3x/5s retry policy would
make a real-SSH-failure test take 15+ seconds.
"""
import time

import pytest

import dashboard.routes.test_runner as test_runner_route
from shared import db as db_module
# Aliased on import -- pytest auto-collects anything matching Test* once
# bound as a name in this module (same reason Story 10.1's test_ssh_pool.py
# aliases the imported test_all_nodes function).
from shared.models import TestRunnerConfig as RunnerConfigModel
from shared.models import TestRunResult as RunResultModel
from worker.executor.ssh_executor import BackgroundCommandHandle
from worker.executor.test_runner import framework as fw


def _login(client):
    client.post("/login", data={"username": "admin", "password": "admin"})


def _configure_nodes(monkeypatch, settings, *, mon="", mgr="", osd="", rgw=""):
    monkeypatch.setattr(settings, "ceph_mon_nodes", mon)
    monkeypatch.setattr(settings, "ceph_mgr_nodes", mgr)
    monkeypatch.setattr(settings, "ceph_osd_nodes", osd)
    monkeypatch.setattr(settings, "ceph_rgw_nodes", rgw)


def _redirect_baseline_dir(monkeypatch, tmp_path):
    """Points the shared baseline-upload directory at a tmp dir instead of
    the real project-root test_runner_baselines/ -- otherwise running this
    suite would write real files into the checkout. Patches
    `shared.test_runner_baselines.BASELINE_FILES_DIR` (via
    `test_runner_route.baselines`, the same module object) rather than a
    module-local name on test_runner_route itself -- the route file only
    ever reads `baselines.BASELINE_FILES_DIR` (qualified), so patching the
    shared module's own attribute is the one place that actually affects
    every reader (this route file AND worker/executor/test_runner/group_b.py)."""
    monkeypatch.setattr(test_runner_route.baselines, "BASELINE_FILES_DIR", tmp_path / "test_runner_baselines")


# -- GET /api/test-runner/config ---------------------------------------------


def test_unauthenticated_get_config_redirects_to_login(dashboard_client):
    response = dashboard_client.get("/api/test-runner/config", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_get_config_empty_state_returns_defaults_not_error(dashboard_client, monkeypatch):
    from config.settings import settings

    _configure_nodes(monkeypatch, settings, mon="10.0.0.1")
    _login(dashboard_client)

    response = dashboard_client.get("/api/test-runner/config")

    assert response.status_code == 200
    data = response.json()
    assert data["nodes"] == [{"host": "10.0.0.1", "roles": ["MON"]}]
    assert data["rgw_endpoint_zone_a"] is None
    assert data["rgw_endpoint_zone_b"] is None
    assert data["rgw_endpoint_vip"] is None
    assert data["test_groups"] == []
    assert data["priorities"] == []
    assert data["baseline_files"] == {key: False for key in test_runner_route.baselines.BASELINE_FILE_KEYS}

    with db_module.SessionLocal() as session:
        assert session.query(RunnerConfigModel).first() is None


# -- POST /api/test-runner/config --------------------------------------------


def test_unauthenticated_post_config_redirects_to_login(dashboard_client):
    response = dashboard_client.post("/api/test-runner/config", json={}, follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_save_then_reload_round_trip(dashboard_client):
    _login(dashboard_client)

    save_response = dashboard_client.post(
        "/api/test-runner/config",
        json={
            "rgw_endpoint_zone_a": "http://rgw-a:7480",
            "rgw_endpoint_zone_b": "http://rgw-b:7480",
            "rgw_endpoint_vip": "http://rgw-vip:7480",
            "test_groups": ["A", "B"],
            "priorities": ["P1"],
        },
    )
    assert save_response.status_code == 200

    reload_response = dashboard_client.get("/api/test-runner/config")
    assert reload_response.status_code == 200
    data = reload_response.json()
    assert data["rgw_endpoint_zone_a"] == "http://rgw-a:7480"
    assert data["rgw_endpoint_zone_b"] == "http://rgw-b:7480"
    assert data["rgw_endpoint_vip"] == "http://rgw-vip:7480"
    assert data["test_groups"] == ["A", "B"]
    assert data["priorities"] == ["P1"]

    # Only one singleton row -- a second save upserts, doesn't add a row.
    with db_module.SessionLocal() as session:
        assert session.query(RunnerConfigModel).count() == 1


def test_unchecking_all_test_groups_persists_empty_not_all_selected(dashboard_client):
    _login(dashboard_client)

    dashboard_client.post(
        "/api/test-runner/config",
        json={"test_groups": ["A", "B", "C", "D"], "priorities": ["P1", "P2"]},
    )
    # Now save again with everything unchecked.
    response = dashboard_client.post(
        "/api/test-runner/config",
        json={"test_groups": [], "priorities": []},
    )
    assert response.status_code == 200

    reload_response = dashboard_client.get("/api/test-runner/config")
    data = reload_response.json()
    assert data["test_groups"] == []
    assert data["priorities"] == []


# -- POST /api/test-runner/config/baseline/{baseline_key} -------------------


def _upload_baseline(client, baseline_key, content=b"abc123", filename="ignored-client-name.txt"):
    return client.post(
        f"/api/test-runner/config/baseline/{baseline_key}",
        files={"file": (filename, content, "text/plain")},
    )


def test_unauthenticated_baseline_upload_redirects_to_login(dashboard_client):
    response = dashboard_client.post(
        "/api/test-runner/config/baseline/rbd_rep.sha256",
        files={"file": ("x.txt", b"abc", "text/plain")},
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_upload_baseline_marks_it_present(dashboard_client, monkeypatch, tmp_path):
    _redirect_baseline_dir(monkeypatch, tmp_path)
    _login(dashboard_client)

    response = _upload_baseline(dashboard_client, "rbd_rep.sha256", content=b"deadbeef  file\n")

    assert response.status_code == 200
    data = response.json()
    assert data["baseline_files"]["rbd_rep.sha256"] is True
    # Every other baseline key still not present.
    assert data["baseline_files"]["cephfs.sha256"] is False

    saved_path = tmp_path / "test_runner_baselines" / "rbd_rep.sha256"
    assert saved_path.read_bytes() == b"deadbeef  file\n"

    # File saved under the fixed baseline_key, not the client's filename.
    assert not (tmp_path / "test_runner_baselines" / "ignored-client-name.txt").exists()

    reload_response = dashboard_client.get("/api/test-runner/config")
    assert reload_response.json()["baseline_files"]["rbd_rep.sha256"] is True


def test_upload_unknown_baseline_key_rejected_with_400(dashboard_client, monkeypatch, tmp_path):
    _redirect_baseline_dir(monkeypatch, tmp_path)
    _login(dashboard_client)

    response = _upload_baseline(dashboard_client, "not_a_real_baseline.txt")

    assert response.status_code == 400
    with db_module.SessionLocal() as session:
        assert session.query(RunnerConfigModel).first() is None


def test_upload_all_seven_baseline_keys_succeed(dashboard_client, monkeypatch, tmp_path):
    _redirect_baseline_dir(monkeypatch, tmp_path)
    _login(dashboard_client)

    for key in test_runner_route.baselines.BASELINE_FILE_KEYS:
        response = _upload_baseline(dashboard_client, key, content=key.encode())
        assert response.status_code == 200

    final = dashboard_client.get("/api/test-runner/config").json()
    assert final["baseline_files"] == {key: True for key in test_runner_route.baselines.BASELINE_FILE_KEYS}


# -- POST /api/test-runner/ssh-check -----------------------------------------


def test_unauthenticated_ssh_check_redirects_to_login(dashboard_client):
    response = dashboard_client.post("/api/test-runner/ssh-check", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_ssh_check_zero_configured_nodes_returns_empty_dict(dashboard_client, monkeypatch):
    from config.settings import settings

    _configure_nodes(monkeypatch, settings)  # all blank
    _login(dashboard_client)

    response = dashboard_client.post("/api/test-runner/ssh-check")

    assert response.status_code == 200
    assert response.json() == {}


def test_ssh_check_mixed_reachability_never_raises(dashboard_client, monkeypatch):
    from config.settings import settings

    _configure_nodes(monkeypatch, settings, mon="host-up", osd="host-down")
    _login(dashboard_client)

    def fake_test_all_nodes(hosts, *args, **kwargs):
        assert set(hosts) == {"host-up", "host-down"}
        return {"host-up": True, "host-down": False}

    monkeypatch.setattr(test_runner_route, "test_all_nodes", fake_test_all_nodes)

    response = dashboard_client.post("/api/test-runner/ssh-check")

    assert response.status_code == 200
    assert response.json() == {"host-up": True, "host-down": False}


def test_ssh_check_all_hosts_unreachable_returns_all_false(dashboard_client, monkeypatch):
    from config.settings import settings

    _configure_nodes(monkeypatch, settings, mon="host-a,host-b")
    _login(dashboard_client)

    def fake_test_all_nodes(hosts, *args, **kwargs):
        return {h: False for h in hosts}

    monkeypatch.setattr(test_runner_route, "test_all_nodes", fake_test_all_nodes)

    response = dashboard_client.post("/api/test-runner/ssh-check")

    assert response.status_code == 200
    assert response.json() == {"host-a": False, "host-b": False}


# -- Story 10.6: run-engine endpoints ----------------------------------------


class _FakeOneShot(fw.TestCase):
    id = "TC-FAKE-ONE"
    name = "fake one-shot"
    group = fw.TestGroup.A
    priority = fw.TestPriority.P1
    background = False

    def run(self, ctx):
        return fw.TestResult(
            test_id=self.id,
            status=fw.TestStatus.RUNNING,
            criteria=[fw.CriterionResult("dummy criterion", passed=True)],
            raw_output="one-shot output\n",
        )


class _FakeErrorOneShot(fw.TestCase):
    id = "TC-FAKE-ERROR"
    name = "fake one-shot that raises"
    group = fw.TestGroup.A
    priority = fw.TestPriority.P2
    background = False

    def run(self, ctx):
        raise ValueError("boom -- a genuine bug in this fake TestCase")


class _FakeBackground(fw.TestCase):
    """poll() resolves on its 2nd call -- lets tests observe an
    intermediate RUNNING poll (accumulating raw_output) before the terminal
    PASS."""

    id = "TC-FAKE-BG"
    name = "fake background"
    group = fw.TestGroup.B
    priority = fw.TestPriority.P1
    background = True

    def start(self, ctx, **kwargs):
        return {"polls": 0}

    def poll(self, ctx, state):
        polls = state["polls"] + 1
        new_state = {"polls": polls}
        if polls < 2:
            criteria = [fw.CriterionResult("still waiting", passed=None, detail=f"poll {polls}")]
        else:
            criteria = [fw.CriterionResult("done", passed=True, detail=f"poll {polls}")]
        result = fw.TestResult(
            test_id=self.id,
            status=fw.TestStatus.RUNNING,
            criteria=criteria,
            raw_output=f"poll-{polls}\n",
        )
        return new_state, result


class _FakeDeclinedStart(fw.TestCase):
    id = "TC-FAKE-DECLINED"
    name = "fake background that declines on start"
    group = fw.TestGroup.C
    priority = fw.TestPriority.P3
    background = True

    def start(self, ctx, **kwargs):
        raise fw.TestCaseDeclined(f"{self.id}: not automatable")


class _FakeBackgroundWithHandle(fw.TestCase):
    """poll() never resolves on its own -- models the real
    effectively-unbounded background loads (TC-RUN-001/010, TC-COMPAT-001,
    TC-PERF-005/007/009) the new POST .../cancel endpoint exists to stop.
    start() returns a REAL BackgroundCommandHandle (so the route's own
    isinstance() guard passes) wrapping a channel that's never actually
    touched -- these tests monkeypatch test_runner_route.cancel_background
    itself rather than exercising real SSH, same posture as this file's
    existing ssh_check tests monkeypatching test_all_nodes."""

    id = "TC-FAKE-BG-HANDLE"
    name = "fake background with a cancellable handle"
    group = fw.TestGroup.D
    priority = fw.TestPriority.P1
    background = True

    def start(self, ctx, **kwargs):
        return {"handle": BackgroundCommandHandle("client1", "fio --name=fake", object())}

    def poll(self, ctx, state):
        result = fw.TestResult(
            test_id=self.id,
            status=fw.TestStatus.RUNNING,
            criteria=[fw.CriterionResult("van dang chay", passed=None)],
            raw_output="still-running\n",
        )
        return state, result


@pytest.fixture
def fake_registry(monkeypatch):
    """Swaps the module-level TEST_CASES_BY_ID/_run_states for an isolated
    set -- qualified monkeypatch.setattr(test_runner_route, ...) so every
    route function (which reads the bare module-global name) sees the fake
    versions, same pattern as this file's existing `test_runner_route.
    baselines` patching. `_run_states` is reset to a fresh dict per test so
    no state leaks between tests via the real module-level singleton."""
    fakes = {
        _FakeOneShot.id: _FakeOneShot(),
        _FakeErrorOneShot.id: _FakeErrorOneShot(),
        _FakeBackground.id: _FakeBackground(),
        _FakeDeclinedStart.id: _FakeDeclinedStart(),
        _FakeBackgroundWithHandle.id: _FakeBackgroundWithHandle(),
    }
    monkeypatch.setattr(test_runner_route, "TEST_CASES_BY_ID", fakes)
    monkeypatch.setattr(test_runner_route, "_run_states", {})
    return fakes


def _wait_until(predicate, timeout=2.0, interval=0.01):
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if predicate():
            return True
        time.sleep(interval)
    return predicate()


def test_list_tests_unauthenticated_redirects_to_login(dashboard_client):
    response = dashboard_client.get("/api/test-runner/tests", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_list_tests_empty_selection_shows_all(dashboard_client, fake_registry):
    _login(dashboard_client)
    response = dashboard_client.get("/api/test-runner/tests")
    assert response.status_code == 200
    ids = {t["id"] for t in response.json()["tests"]}
    assert ids == set(fake_registry.keys())
    not_started = [t for t in response.json()["tests"] if t["id"] == _FakeOneShot.id][0]
    assert not_started["status"] == "not_started"
    assert not_started["overridden"] is False


def test_list_tests_narrows_to_saved_group_selection(dashboard_client, fake_registry):
    _login(dashboard_client)
    dashboard_client.post(
        "/api/test-runner/config",
        json={"test_groups": ["B"], "priorities": []},
    )
    response = dashboard_client.get("/api/test-runner/tests")
    ids = {t["id"] for t in response.json()["tests"]}
    assert ids == {_FakeBackground.id}


def test_run_unknown_id_returns_404(dashboard_client, fake_registry):
    _login(dashboard_client)
    response = dashboard_client.post("/api/test-runner/tests/TC-DOES-NOT-EXIST/run")
    assert response.status_code == 404


def test_run_one_shot_completes_to_pass(dashboard_client, fake_registry):
    _login(dashboard_client)

    run_response = dashboard_client.post(f"/api/test-runner/tests/{_FakeOneShot.id}/run")
    assert run_response.status_code == 200
    # The fake's run() is instant (no real I/O) -- the daemon thread may
    # already have finished by the time this response is serialized, so
    # either RUNNING (still in flight) or the final PASS is a valid
    # immediate response; what matters is it's never left un-started.
    assert run_response.json()["status"] in (fw.TestStatus.RUNNING.value, fw.TestStatus.PASS.value)

    def _finished():
        r = dashboard_client.get(f"/api/test-runner/tests/{_FakeOneShot.id}/result")
        return r.json()["status"] != fw.TestStatus.RUNNING.value

    assert _wait_until(_finished), "one-shot fake test never left RUNNING"

    final = dashboard_client.get(f"/api/test-runner/tests/{_FakeOneShot.id}/result").json()
    assert final["status"] == fw.TestStatus.PASS.value
    assert final["raw_output"] == "one-shot output\n"
    assert final["criteria"] == [{"description": "dummy criterion", "passed": True, "detail": ""}]


def test_run_one_shot_uncaught_exception_maps_to_error(dashboard_client, fake_registry):
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeErrorOneShot.id}/run")

    def _finished():
        r = dashboard_client.get(f"/api/test-runner/tests/{_FakeErrorOneShot.id}/result")
        return r.json()["status"] != fw.TestStatus.RUNNING.value

    assert _wait_until(_finished), "fake error test never left RUNNING"
    final = dashboard_client.get(f"/api/test-runner/tests/{_FakeErrorOneShot.id}/result").json()
    assert final["status"] == fw.TestStatus.ERROR.value
    assert "boom" in final["notes"]


def test_run_twice_while_running_returns_409(dashboard_client, fake_registry):
    _login(dashboard_client)
    first = dashboard_client.post(f"/api/test-runner/tests/{_FakeBackground.id}/run")
    assert first.status_code == 200
    second = dashboard_client.post(f"/api/test-runner/tests/{_FakeBackground.id}/run")
    assert second.status_code == 409


def test_background_start_declined_maps_to_skip(dashboard_client, fake_registry):
    _login(dashboard_client)
    response = dashboard_client.post(f"/api/test-runner/tests/{_FakeDeclinedStart.id}/run")
    assert response.status_code == 200
    assert response.json()["status"] == fw.TestStatus.SKIP.value

    result = dashboard_client.get(f"/api/test-runner/tests/{_FakeDeclinedStart.id}/result").json()
    assert result["status"] == fw.TestStatus.SKIP.value


def test_background_poll_accumulates_output_and_resolves(dashboard_client, fake_registry):
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeBackground.id}/run")

    first_poll = dashboard_client.get(f"/api/test-runner/tests/{_FakeBackground.id}/result").json()
    assert first_poll["status"] == fw.TestStatus.RUNNING.value
    assert first_poll["raw_output"] == "poll-1\n"

    second_poll = dashboard_client.get(f"/api/test-runner/tests/{_FakeBackground.id}/result").json()
    assert second_poll["status"] == fw.TestStatus.PASS.value
    # accumulated across both polls, not overwritten by the 2nd delta alone
    assert second_poll["raw_output"] == "poll-1\npoll-2\n"

    # terminal -- a 3rd GET must not poll again (no 3rd delta appended)
    third_poll = dashboard_client.get(f"/api/test-runner/tests/{_FakeBackground.id}/result").json()
    assert third_poll["raw_output"] == "poll-1\npoll-2\n"


# -- POST /api/test-runner/tests/{id}/cancel ---------------------------------
# 2026-08-07: TC-RUN-001/010, TC-COMPAT-001, TC-PERF-005/007/009 (and their
# Group E S3 equivalents) all launch effectively-unbounded remote I/O load
# via execute_background() -- before this endpoint existed, this app had
# NO way to stop one short of an operator manually SSHing in and killing
# processes by hand, which on a small/lab cluster can run the CPU/RAM up
# until the cluster itself falls over. These tests monkeypatch
# ssh_executor.cancel_background() (imported into test_runner_route's own
# namespace) rather than exercising real SSH -- same posture as this
# file's ssh_check tests monkeypatching test_all_nodes.
#
# 2026-08-07, extended same day (incident follow-up): the original version
# of this endpoint 409'd whenever `_run_states` had no live RUNNING entry
# for the test_id -- which is EXACTLY the situation after a Dashboard
# restart (that dict is in-memory only, see its own docstring), silently
# leaving an operator with no way to stop a still-running remote process
# via the UI (the frontend's "Hủy" button was also only rendered while
# status=="running", so it was invisible too -- a real incident: the app
# kept eating a real Ceph cluster's RAM with no visible off switch). The
# endpoint now NEVER 409s for this reason -- it always attempts a
# pattern-based `pkill -f` fallback (test_runner_route._pattern_kill(),
# monkeypatched to a fake here rather than exercising real SSH) in addition
# to the PGID-based kill, and reports success either way.


def test_cancel_unknown_id_returns_404(dashboard_client, fake_registry):
    _login(dashboard_client)
    response = dashboard_client.post("/api/test-runner/tests/TC-DOES-NOT-EXIST/cancel")
    assert response.status_code == 404


def test_cancel_never_started_falls_back_to_pattern_kill(dashboard_client, fake_registry, monkeypatch):
    """No /run was ever called -- `_run_states` has no entry at all for
    this test_id (models a Dashboard restart just as well as an actual
    restart would: either way, the tracked handle is gone). Must still
    succeed via the pattern-kill fallback rather than 409ing."""
    _login(dashboard_client)
    pattern_calls = []
    monkeypatch.setattr(
        test_runner_route,
        "_pattern_kill",
        lambda test_id: pattern_calls.append(test_id) or [f"fakehost: đã kill tiến trình khớp pattern"],
    )

    response = dashboard_client.post(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/cancel")

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == fw.TestStatus.FAIL.value
    assert body["overridden"] is True
    assert "Dashboard không có bản ghi" in body["override_note"]
    assert "fakehost" in body["override_note"]
    assert pattern_calls == [_FakeBackgroundWithHandle.id]


def test_cancel_kills_remote_process_and_marks_terminal(dashboard_client, fake_registry, monkeypatch):
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/run")

    calls = []
    monkeypatch.setattr(
        test_runner_route, "cancel_background", lambda handle: calls.append(handle) or True
    )

    response = dashboard_client.post(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/cancel")

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == fw.TestStatus.FAIL.value
    assert body["overridden"] is True
    assert "Đã gửi lệnh kill theo PGID" in body["override_note"]
    assert len(calls) == 1
    assert isinstance(calls[0], BackgroundCommandHandle)

    # Sticky against further polling -- get_test_result() must not resume
    # RUNNING or overwrite the cancel note (same guard override uses).
    result = dashboard_client.get(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/result").json()
    assert result["status"] == fw.TestStatus.FAIL.value
    assert "Đã gửi lệnh kill theo PGID" in result["override_note"]


def test_cancel_reports_unconfirmed_when_no_pgid_was_ever_captured(
    dashboard_client, fake_registry, monkeypatch
):
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/run")
    monkeypatch.setattr(test_runner_route, "cancel_background", lambda handle: False)

    response = dashboard_client.post(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/cancel")

    assert response.status_code == 200
    body = response.json()
    # Still marked terminal (so polling stops) even though the kill itself
    # couldn't be confirmed -- the note must say so rather than falsely
    # claiming success.
    assert body["status"] == fw.TestStatus.FAIL.value
    assert "KHÔNG xác nhận được" in body["override_note"]


def test_cancel_twice_stays_idempotent_and_succeeds_both_times(dashboard_client, fake_registry, monkeypatch):
    """Deliberately NOT a 409 the second time (unlike the old design) --
    an operator unsure whether a previous cancel actually reached the
    remote host must be able to just click "Hủy" again. cancel_test()
    re-reads the still-live handle from `_run_states` (cancel_test never
    clears background_state) and re-attempts both kill paths -- calling
    cancel_background twice is itself idempotent (kill -TERM/-KILL on an
    already-dead process group is a harmless no-op)."""
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/run")
    calls = []
    monkeypatch.setattr(
        test_runner_route, "cancel_background", lambda handle: calls.append(handle) or True
    )
    monkeypatch.setattr(test_runner_route, "_pattern_kill", lambda test_id: [])

    first = dashboard_client.post(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/cancel")
    assert first.status_code == 200
    second = dashboard_client.post(f"/api/test-runner/tests/{_FakeBackgroundWithHandle.id}/cancel")
    assert second.status_code == 200
    assert len(calls) == 2


def test_cancel_one_shot_test_returns_409(dashboard_client, fake_registry):
    # Cancel only makes sense for a background test's live SSH handle -- a
    # one-shot test is rejected outright regardless of its current status.
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeOneShot.id}/run")
    response = dashboard_client.post(f"/api/test-runner/tests/{_FakeOneShot.id}/cancel")
    assert response.status_code == 409


# -- _pattern_kill() / _kill_target_hosts() -----------------------------------
# Unit-level coverage of the pattern-fallback machinery itself (not routed
# through fake_registry -- these exercise the REAL _BACKGROUND_KILL_PATTERNS
# table against a real test_id), monkeypatching execute_command (imported
# into test_runner_route's own namespace) rather than exercising real SSH,
# same posture as the cancel_background tests above.


def test_pattern_kill_unregistered_test_id_is_a_noop(monkeypatch):
    calls = []
    monkeypatch.setattr(test_runner_route, "execute_command", lambda host, cmd: calls.append((host, cmd)))

    lines = test_runner_route._pattern_kill("TC-NOT-A-BACKGROUND-TEST")

    assert lines == []
    assert calls == []


def test_pattern_kill_client_host_target_uses_saved_config(dashboard_client, monkeypatch):
    _login(dashboard_client)
    dashboard_client.post("/api/test-runner/config", json={"client_host": "10.0.0.9"})

    calls = []

    def fake_execute_command(host, cmd):
        calls.append((host, cmd))
        return "KILLED\n"

    monkeypatch.setattr(test_runner_route, "execute_command", fake_execute_command)

    lines = test_runner_route._pattern_kill("TC-RUN-001")

    assert calls == [("10.0.0.9", "pkill -f 'fio --name=upgrade_io_test' && echo KILLED || echo NONE")]
    assert lines == ["10.0.0.9: đã kill tiến trình khớp pattern đã biết"]


def test_pattern_kill_no_client_host_configured_targets_nothing(dashboard_client):
    # No /config POST in this test -- TestRunnerConfig row doesn't even
    # exist yet, same as a fresh install nobody has configured.
    lines = test_runner_route._pattern_kill("TC-RUN-001")
    assert lines == []


def test_pattern_kill_osd_target_checks_every_configured_osd_host(monkeypatch):
    from config.settings import settings

    _configure_nodes(monkeypatch, settings, osd="osd1,osd2")
    monkeypatch.setattr(test_runner_route, "execute_command", lambda host, cmd: "NONE\n")

    lines = test_runner_route._pattern_kill("TC-RUN-013")

    assert lines == [
        "osd1: không thấy tiến trình nào khớp pattern",
        "osd2: không thấy tiến trình nào khớp pattern",
    ]


def test_pattern_kill_reports_ssh_failure_per_host_without_raising(dashboard_client, monkeypatch):
    from worker.executor.ssh_executor import ExecutorError

    _login(dashboard_client)
    dashboard_client.post("/api/test-runner/config", json={"client_host": "unreachable-host"})

    def raising_execute_command(host, cmd):
        raise ExecutorError(f"{host}: connection refused")

    monkeypatch.setattr(test_runner_route, "execute_command", raising_execute_command)

    lines = test_runner_route._pattern_kill("TC-RUN-001")

    assert len(lines) == 1
    assert "lỗi SSH khi kill" in lines[0]


def test_override_unknown_id_returns_404(dashboard_client, fake_registry):
    _login(dashboard_client)
    response = dashboard_client.post(
        "/api/test-runner/tests/TC-DOES-NOT-EXIST/override", json={"status": "pass"}
    )
    assert response.status_code == 404


def test_override_invalid_status_returns_400(dashboard_client, fake_registry):
    _login(dashboard_client)
    response = dashboard_client.post(
        f"/api/test-runner/tests/{_FakeBackground.id}/override", json={"status": "maybe"}
    )
    assert response.status_code == 400


def test_override_is_sticky_against_further_polling(dashboard_client, fake_registry):
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeBackground.id}/run")
    # one real poll first, so there's something to overwrite if the sticky
    # guard didn't work
    dashboard_client.get(f"/api/test-runner/tests/{_FakeBackground.id}/result")

    override_response = dashboard_client.post(
        f"/api/test-runner/tests/{_FakeBackground.id}/override",
        json={"status": "fail", "note": "vận hành xác nhận thất bại thủ công"},
    )
    assert override_response.status_code == 200
    assert override_response.json()["status"] == fw.TestStatus.FAIL.value
    assert override_response.json()["overridden"] is True

    # a further GET must NOT poll again and must not flip status back
    after = dashboard_client.get(f"/api/test-runner/tests/{_FakeBackground.id}/result").json()
    assert after["status"] == fw.TestStatus.FAIL.value
    assert after["overridden"] is True
    assert after["override_note"] == "vận hành xác nhận thất bại thủ công"
    assert after["raw_output"] == "poll-1\n"  # unchanged since the override


# -- Story 10.7: report export endpoints -------------------------------------
#
# Deliberately does NOT reuse the `fake_registry` fixture above -- its
# "TC-FAKE-*" ids don't match any of the 4 real document-id prefixes
# (TC-RUN-/TC-POST-/TC-COMPAT-/TC-PERF-) worker/executor/test_runner/
# report.py's _group_for_doc_id() recognizes, so it would raise. Report
# generation does no SSH (see report.py's own module docstring), so testing
# against the REAL TEST_CASES_BY_ID (67 classes, always importable) is both
# simpler and more representative -- only `_run_states` needs faking.


def test_report_markdown_unauthenticated_redirects_to_login(dashboard_client):
    response = dashboard_client.get("/api/test-runner/report/markdown", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_report_excel_unauthenticated_redirects_to_login(dashboard_client):
    response = dashboard_client.get("/api/test-runner/report/excel", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_report_summary_unauthenticated_redirects_to_login(dashboard_client):
    response = dashboard_client.get("/api/test-runner/report/summary", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_report_markdown_contains_known_rows_and_attachment_headers(dashboard_client, monkeypatch):
    monkeypatch.setattr(
        test_runner_route,
        "_run_states",
        {
            "TC-POST-001": test_runner_route._RunState(
                status=fw.TestStatus.PASS.value,
                criteria=[{"description": "x", "passed": True, "detail": ""}],
                finished_at="2026-08-05T10:00:00",
            )
        },
    )
    _login(dashboard_client)

    response = dashboard_client.get("/api/test-runner/report/markdown")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/markdown")
    assert 'attachment; filename="bao-cao-nang-cap-ceph.md"' in response.headers["content-disposition"]
    body = response.text
    assert "TC-POST-001" in body
    assert "TC-RUN-001" in body  # a never-run test still gets a row
    assert "Bảng tổng hợp" in body


def test_report_excel_round_trips_and_has_attachment_headers(dashboard_client, monkeypatch):
    import io

    from openpyxl import load_workbook

    monkeypatch.setattr(test_runner_route, "_run_states", {})
    _login(dashboard_client)

    response = dashboard_client.get("/api/test-runner/report/excel")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'attachment; filename="bao-cao-nang-cap-ceph.xlsx"' in response.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == ["Ket qua test case", "Tong hop"]


def test_report_summary_returns_json_with_totals(dashboard_client, monkeypatch):
    monkeypatch.setattr(test_runner_route, "_run_states", {})
    _login(dashboard_client)

    response = dashboard_client.get("/api/test-runner/report/summary")

    assert response.status_code == 200
    data = response.json()
    assert "summary_text" in data
    assert "PASS" in data["summary_text"]


def test_report_reflects_override_and_run013_osd_subtable(dashboard_client, monkeypatch):
    monkeypatch.setattr(
        test_runner_route,
        "_run_states",
        {
            "TC-POST-002": test_runner_route._RunState(
                status=fw.TestStatus.FAIL.value,
                overridden=True,
                override_note="da xac nhan HEALTH_OK bang tay",
            ),
            "TC-RUN-013": test_runner_route._RunState(
                status=fw.TestStatus.PENDING.value,
                background_state={
                    "completed": [{"osd_id": 5, "seconds": 12.3, "exit_code": 0, "over_estimate": False}]
                },
            ),
        },
    )
    _login(dashboard_client)

    body = dashboard_client.get("/api/test-runner/report/markdown").text

    assert "[Override] da xac nhan HEALTH_OK bang tay" in body
    assert "Chi tiết TC-RUN-013" in body
    assert "| 5 | 12.3 | 0 | Không |" in body


# -----------------------------------------------------------------------
# Story 10.8: results survive a Dashboard restart. `_persist_run_state()`/
# `_load_persisted_run_states()` bridge `_run_states` <-> `RunResultModel`;
# `POST /reset` is the new explicit "start a new campaign" endpoint this
# story's own persistence work necessitated (a restart used to be the
# de-facto reset -- it deliberately no longer is one).
# -----------------------------------------------------------------------


def test_run_one_shot_auto_saves_terminal_result_to_db(dashboard_client, fake_registry):
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeOneShot.id}/run")

    # Wait on the DB row itself, not just _run_states via the HTTP endpoint:
    # _apply_result() flips the in-memory status BEFORE the auto-save DB
    # write happens (persist runs after releasing _run_lock) -- polling the
    # HTTP endpoint alone would be a TOCTOU race against the still-in-flight
    # daemon-thread DB commit.
    def _row_saved():
        with db_module.SessionLocal() as session:
            row = session.get(RunResultModel, _FakeOneShot.id)
            return row is not None and row.status != fw.TestStatus.RUNNING.value

    assert _wait_until(_row_saved), "one-shot fake test's result never auto-saved to DB"

    with db_module.SessionLocal() as session:
        row = session.get(RunResultModel, _FakeOneShot.id)
        assert row is not None
        assert row.status == fw.TestStatus.PASS.value
        assert row.raw_output == "one-shot output\n"
        assert row.overridden is False


def test_override_auto_saves_to_db(dashboard_client, fake_registry):
    _login(dashboard_client)
    dashboard_client.post(f"/api/test-runner/tests/{_FakeOneShot.id}/run")
    dashboard_client.post(
        f"/api/test-runner/tests/{_FakeOneShot.id}/override",
        json={"status": "fail", "note": "ghi de thu cong"},
    )

    with db_module.SessionLocal() as session:
        row = session.get(RunResultModel, _FakeOneShot.id)
        assert row is not None
        assert row.status == fw.TestStatus.FAIL.value
        assert row.overridden is True
        assert row.override_note == "ghi de thu cong"


def test_load_persisted_run_states_restores_terminal_and_overridden_rows(dashboard_client):
    _login(dashboard_client)
    with db_module.SessionLocal() as session:
        session.add(
            RunResultModel(
                test_id="TC-FAKE-TERMINAL",
                status=fw.TestStatus.PASS.value,
                criteria_json='[{"description": "x", "passed": true, "detail": ""}]',
                raw_output="done\n",
                notes="",
                overridden=False,
            )
        )
        session.add(
            RunResultModel(
                test_id="TC-FAKE-OVERRIDDEN",
                status=fw.TestStatus.FAIL.value,
                overridden=True,
                override_note="manual call",
            )
        )
        session.commit()

    test_runner_route._run_states.clear()
    test_runner_route._load_persisted_run_states()

    restored_terminal = test_runner_route._run_states["TC-FAKE-TERMINAL"]
    assert restored_terminal.status == fw.TestStatus.PASS.value
    assert restored_terminal.raw_output == "done\n"
    assert restored_terminal.criteria == [{"description": "x", "passed": True, "detail": ""}]

    restored_overridden = test_runner_route._run_states["TC-FAKE-OVERRIDDEN"]
    assert restored_overridden.status == fw.TestStatus.FAIL.value
    assert restored_overridden.overridden is True
    assert restored_overridden.override_note == "manual call"


def test_load_persisted_run_states_normalizes_orphaned_running_row_to_error(dashboard_client):
    _login(dashboard_client)
    with db_module.SessionLocal() as session:
        session.add(
            RunResultModel(
                test_id="TC-FAKE-ORPHANED",
                status=fw.TestStatus.RUNNING.value,
                overridden=False,
            )
        )
        session.commit()

    test_runner_route._run_states.clear()
    test_runner_route._load_persisted_run_states()

    restored = test_runner_route._run_states["TC-FAKE-ORPHANED"]
    assert restored.status == fw.TestStatus.ERROR.value
    assert restored.finished_at is not None
    assert "khởi động lại" in restored.notes
    assert restored.background_state is None  # never attempted to restore/fabricate one

    # The correction must be re-persisted too -- the DB must not keep
    # claiming "running" forever just because only the in-memory copy was
    # fixed.
    with db_module.SessionLocal() as session:
        row = session.get(RunResultModel, "TC-FAKE-ORPHANED")
        assert row.status == fw.TestStatus.ERROR.value


def test_load_persisted_run_states_leaves_overridden_running_alone(dashboard_client):
    """An overridden test is, by definition, already closed out by an
    operator -- even if its stored status somehow says RUNNING (shouldn't
    happen via this app's own routes, but the normalization guard is keyed
    on overridden specifically, not just status, so this locks that in)."""
    with db_module.SessionLocal() as session:
        session.add(
            RunResultModel(
                test_id="TC-FAKE-OVERRIDDEN-RUNNING",
                status=fw.TestStatus.RUNNING.value,
                overridden=True,
                override_note="closed by hand",
            )
        )
        session.commit()

    test_runner_route._run_states.clear()
    test_runner_route._load_persisted_run_states()

    restored = test_runner_route._run_states["TC-FAKE-OVERRIDDEN-RUNNING"]
    assert restored.status == fw.TestStatus.RUNNING.value
    assert restored.overridden is True


# -- POST /api/test-runner/reset -----------------------------------------


def test_unauthenticated_reset_redirects_to_login(dashboard_client):
    response = dashboard_client.post("/api/test-runner/reset", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_reset_clears_run_states_and_db_but_not_config(dashboard_client, fake_registry):
    _login(dashboard_client)
    dashboard_client.post(
        "/api/test-runner/config",
        json={"rgw_endpoint_zone_a": "https://rgw-a.example", "test_groups": ["A"]},
    )
    dashboard_client.post(f"/api/test-runner/tests/{_FakeOneShot.id}/run")

    # Wait for the daemon thread's OWN auto-save to actually land before
    # resetting -- otherwise a still-in-flight background persist could
    # write its row back in right after /reset's DELETE runs (same TOCTOU
    # class as test_run_one_shot_auto_saves_terminal_result_to_db above).
    def _row_saved():
        with db_module.SessionLocal() as session:
            row = session.get(RunResultModel, _FakeOneShot.id)
            return row is not None and row.status != fw.TestStatus.RUNNING.value

    assert _wait_until(_row_saved), "one-shot fake test's result never auto-saved to DB"

    response = dashboard_client.post("/api/test-runner/reset")
    assert response.status_code == 200

    assert test_runner_route._run_states == {}
    with db_module.SessionLocal() as session:
        assert session.query(RunResultModel).count() == 0
        config = session.query(RunnerConfigModel).first()
        assert config is not None
        assert config.rgw_endpoint_zone_a == "https://rgw-a.example"

    after_reset = dashboard_client.get(f"/api/test-runner/tests/{_FakeOneShot.id}/result").json()
    assert after_reset["status"] == "not_started"
